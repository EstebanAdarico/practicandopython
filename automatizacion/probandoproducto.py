import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import pandas as pd
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================
HEADER_SCAN_ROWS = 30  # busca encabezados en las primeras 30 filas

# =========================
# UTILIDADES
# =========================
def safe_str(x):
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s.lower() == "nan" else s

def safe_int(x):
    try:
        s = safe_str(x).replace(",", "")
        if s == "":
            return None
        return int(float(s))
    except Exception:
        return None

def find_header_row(df_preview, header_name="DOC"):
    target = header_name.strip().upper()
    for r in range(len(df_preview)):
        row = df_preview.iloc[r].astype(str).str.strip().str.upper()
        if target in set(row.values):
            return r
    return None

def col_pick(df, candidates):
    """
    Devuelve el nombre real de columna que matchea cualquiera de candidates.
    la ptm dos horas en esto no modificar hasta nuevo aviso
    """
    cols_upper = {c.upper().strip(): c for c in df.columns}
    for cand in candidates:
        cand = cand.upper().strip()
        if cand in cols_upper:
            return cols_upper[cand]
    return None

def normalize_name(s: str) -> str:
    return "".join(ch for ch in s.upper() if ch.isalnum())

def detectar_hoja_diarios(excel_path: str) -> str:
    """
    Detecta automáticamente la hoja DIARIO_CONSUMO / DIARIOS_CONSUMO aunque tenga espacios.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()

    # 1) preferir hoja que contenga DIARIO(S) + CONSUMO
    for name in names:
        n = normalize_name(name)
        if ("DIARIO" in n or "DIARIOS" in n) and "CONSUMO" in n:
            return name

    raise ValueError(f"No encontré una hoja tipo DIARIO(S)_CONSUMO. Hojas: {names}")

def read_diarios(excel_path):
    """
    Lee la hoja DIARIO(S)_CONSUMO detectando header automáticamente.
    Devuelve df, meta, sheet_real.
    """
    sheet_real = detectar_hoja_diarios(excel_path)

    preview = pd.read_excel(
        excel_path,
        sheet_name=sheet_real,
        header=None,
        nrows=HEADER_SCAN_ROWS,
        dtype=str,
        engine="openpyxl"
    )
    hdr = find_header_row(preview, "DOC")
    if hdr is None:
        raise ValueError(f"No pude detectar encabezado 'DOC' en la hoja {sheet_real}.")

    df = pd.read_excel(
        excel_path,
        sheet_name=sheet_real,
        header=hdr,
        dtype=str,
        engine="openpyxl"
    )

    # Detectar columnas (según tu imagen)
    col_doc     = col_pick(df, ["DOC"])
    col_fecha   = col_pick(df, ["FECHA"])
    col_cliente = col_pick(df, ["CLIENTE"])
    col_reserva = col_pick(df, ["RESERVA"])
    col_gr      = col_pick(df, ["G/R SALIDA", "GR SALIDA", "G/R", "GUIA"])
    col_vs      = col_pick(df, ["V/S", "VS"])
    col_canje   = col_pick(df, ["CANJE"])
    col_cant    = col_pick(df, ["SUMA DE S", "SUMA S", "S", "CANTIDAD"])

    if not col_doc:
        raise ValueError(f"No existe columna DOC en {sheet_real}. Columnas: {list(df.columns)}")
    if not col_canje:
        raise ValueError(f"No existe columna CANJE en {sheet_real}. Columnas: {list(df.columns)}")
    if not col_cant:
        raise ValueError(f"No existe columna cantidad (Suma de S / S) en {sheet_real}. Columnas: {list(df.columns)}")

    meta = {
        "DOC": col_doc,
        "FECHA": col_fecha,
        "CLIENTE": col_cliente,
        "RESERVA": col_reserva,
        "GR": col_gr,
        "VS": col_vs,
        "CANJE": col_canje,
        "CANT": col_cant,
    }
    return df, meta, sheet_real

def docs_in_visual_order(df, col_doc):
    """
    DOCs únicos por orden de aparición.
    """
    docs = []
    seen = set()
    for v in df[col_doc].tolist():
        d = safe_int(v)
        if d is None:
            continue
        if d not in seen:
            seen.add(d)
            docs.append(d)
    return docs

def build_rows_for_range(df, meta, doc_ini, doc_fin):
    """
    Devuelve filas para tabla:
    (DOC, ITEM, FECHA, CLIENTE, RESERVA, GR, VS, CANJE, CANT)
    """
    col_doc = meta["DOC"]
    docs = docs_in_visual_order(df, col_doc)
    docs_rango = [d for d in docs if doc_ini <= d <= doc_fin]

    if not docs_rango:
        return [], docs, []

    item_counter = {d: 0 for d in docs_rango}
    out = []

    for _, row in df.iterrows():
        d = safe_int(row.get(col_doc))
        if d is None or d not in item_counter:
            continue

        item_counter[d] += 1

        fecha   = safe_str(row.get(meta["FECHA"])) if meta["FECHA"] else ""
        cliente = safe_str(row.get(meta["CLIENTE"])) if meta["CLIENTE"] else ""
        reserva = safe_str(row.get(meta["RESERVA"])) if meta["RESERVA"] else ""
        gr      = safe_str(row.get(meta["GR"])) if meta["GR"] else ""
        vs      = safe_str(row.get(meta["VS"])) if meta["VS"] else ""
        canje   = safe_str(row.get(meta["CANJE"])) if meta["CANJE"] else ""
        cant    = safe_str(row.get(meta["CANT"])) if meta["CANT"] else ""

        if canje == "" and cant == "":
            continue

        out.append((d, item_counter[d], fecha, cliente, reserva, gr, vs, canje, cant))

    return out, docs, docs_rango


# =========================
# VENTOR (NO TOCAR POR AHORA)
# =========================
def ventor_iniciar_guia(*args, **kwargs):
    # TODO: aquí irá tu lógica pyautogui para Ventor
    pass

def ventor_ingresar_item(*args, **kwargs):
    # TODO: aquí irá tu lógica pyautogui para Ventor
    pass

def ventor_guardar(*args, **kwargs):
    # TODO: aquí irá tu lógica pyautogui para Ventor
    pass


# =========================
# APP
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Antapaccay – DIARIO(S)_CONSUMO → (Ventor luego)")
        self.geometry("1250x720")

        self.excel_path = tk.StringVar()
        self.doc_inicio = tk.StringVar()
        self.doc_fin = tk.StringVar()

        outer = ttk.Frame(self, padding=12)
        outer.pack(fill="both", expand=True)

        # Archivo Excel (mantener)
        row1 = ttk.Frame(outer)
        row1.pack(fill="x")
        ttk.Label(row1, text="Archivo Excel:").pack(side="left")
        ttk.Entry(row1, textvariable=self.excel_path).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(row1, text="Buscar...", command=self.pick_excel).pack(side="left")

        # Rango
        row2 = ttk.Frame(outer)
        row2.pack(fill="x", pady=8)
        ttk.Label(row2, text="DOC inicio:").pack(side="left")
        ttk.Entry(row2, textvariable=self.doc_inicio, width=12).pack(side="left", padx=8)
        ttk.Label(row2, text="DOC fin:").pack(side="left")
        ttk.Entry(row2, textvariable=self.doc_fin, width=12).pack(side="left", padx=8)

        # Botón MOSTRAR
        row3 = ttk.Frame(outer)
        row3.pack(fill="x", pady=8)
        ttk.Button(row3, text="MOSTRAR (cargar desde DIARIO(S)_CONSUMO)", command=self.run_mostrar).pack(side="left")

        # Tabla
        ttk.Label(outer, text="Items del rango (DIARIO(S)_CONSUMO):").pack(anchor="w", pady=(12, 4))

        columns = ("DOC", "ITEM", "FECHA", "CLIENTE", "RESERVA", "G/R SALIDA", "V/S", "CANJE", "CANT")
        self.tree = ttk.Treeview(outer, columns=columns, show="headings", height=16)

        for c in columns:
            self.tree.heading(c, text=c)
            if c in ("DOC", "ITEM"):
                self.tree.column(c, width=70, anchor="center")
            elif c == "CANT":
                self.tree.column(c, width=90, anchor="center")
            elif c in ("FECHA", "V/S"):
                self.tree.column(c, width=90, anchor="center")
            elif c in ("RESERVA", "G/R SALIDA"):
                self.tree.column(c, width=120, anchor="center")
            else:
                self.tree.column(c, width=280, anchor="w")

        self.tree.pack(fill="both", expand=False)

        # Log
        ttk.Label(outer, text="Log:").pack(anchor="w", pady=(12, 4))
        self.log_box = tk.Text(outer, height=10)
        self.log_box.pack(fill="both", expand=True)

    def log(self, msg):
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")

    def pick_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if path:
            self.excel_path.set(path)

    def clear_table(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

    def run_mostrar(self):
        def job():
            try:
                if not self.excel_path.get():
                    messagebox.showwarning("Error", "Selecciona el archivo Excel.")
                    return

                d_ini = safe_int(self.doc_inicio.get())
                d_fin = safe_int(self.doc_fin.get())
                if d_ini is None or d_fin is None:
                    messagebox.showwarning("Error", "DOC inicio/fin deben ser números (ej: 3628).")
                    return
                if d_ini > d_fin:
                    messagebox.showwarning("Error", "DOC inicio debe ser <= DOC fin.")
                    return

                self.log("📌 Leyendo DIARIO(S)_CONSUMO...")
                df, meta, sheet_real = read_diarios(self.excel_path.get())
                self.log(f"✅ Hoja usada: {sheet_real}")

                rows, docs_all, docs_rango = build_rows_for_range(df, meta, d_ini, d_fin)

                self.clear_table()

                if not rows:
                    self.log("⚠️ No se encontraron datos en ese rango.")
                    if docs_all:
                        self.log(f"Primer DOC detectado: {docs_all[0]} | Último: {docs_all[-1]}")
                    return

                for row in rows:
                    self.tree.insert("", "end", values=row)

                self.log(f"✅ DOCs detectados totales: {len(docs_all)}")
                self.log(f"✅ DOCs incluidos en rango: {docs_rango}")
                self.log(f"✅ Filas (items) mostradas: {len(rows)}")

            except Exception as e:
                self.log(f"❌ Error: {e}")

        threading.Thread(target=job, daemon=True).start()


if __name__ == "__main__":
    App().mainloop()
    